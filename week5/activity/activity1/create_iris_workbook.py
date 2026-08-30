"""Create a formatted Excel workbook from the supplied UCI Iris data."""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HERE = Path(__file__).resolve().parent
SOURCE = HERE.parent / "iris" / "iris.data"
OUTPUT = HERE / "iris_dataset.xlsx"
HEADERS = [
    "Sepal length (cm)",
    "Sepal width (cm)",
    "Petal length (cm)",
    "Petal width (cm)",
    "Species",
]


def read_data() -> list[list[object]]:
    rows: list[list[object]] = []
    with SOURCE.open(newline="", encoding="utf-8") as file:
        for source_row in csv.reader(file):
            if not source_row:
                continue
            if len(source_row) != 5:
                raise ValueError(f"Invalid source row: {source_row}")
            rows.append(
                [*[float(value) for value in source_row[:4]], source_row[4].removeprefix("Iris-").title()]
            )
    return rows


def build_workbook(rows: list[list[object]]) -> None:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Iris Data"
    data_sheet.append(HEADERS)
    for row in rows:
        data_sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in data_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:E{len(rows) + 1}"
    data_sheet.column_dimensions["A"].width = 20
    data_sheet.column_dimensions["B"].width = 19
    data_sheet.column_dimensions["C"].width = 20
    data_sheet.column_dimensions["D"].width = 19
    data_sheet.column_dimensions["E"].width = 16
    for row in data_sheet.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.number_format = "0.0"

    table = Table(displayName="IrisDataset", ref=f"A1:E{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    data_sheet.add_table(table)

    summary = workbook.create_sheet("Summary")
    summary.sheet_view.showGridLines = False
    summary["A1"] = "Iris Dataset Summary"
    summary["A1"].font = Font(size=20, bold=True, color="1F4E78")
    summary.merge_cells("A1:D1")

    counts = Counter(str(row[4]) for row in rows)
    summary_rows = [
        ("Source", "UCI Iris Plants Database (iris.data)"),
        ("Observations", len(rows)),
        ("Numeric features", 4),
        ("Missing values", 0),
        ("Setosa", counts["Setosa"]),
        ("Versicolor", counts["Versicolor"]),
        ("Virginica", counts["Virginica"]),
    ]
    for index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(index, 1, label)
        summary.cell(index, 2, value)
        summary.cell(index, 1).font = Font(bold=True, color="FFFFFF")
        summary.cell(index, 1).fill = PatternFill("solid", fgColor="5B9BD5")
        summary.cell(index, 2).fill = PatternFill("solid", fgColor="DDEBF7")
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 40

    chart = ScatterChart()
    chart.title = "Petal measurements by species"
    chart.x_axis.title = "Petal length (cm)"
    chart.y_axis.title = "Petal width (cm)"
    chart.height = 10
    chart.width = 17
    chart.legend.position = "r"

    # The source file is ordered in three contiguous groups of 50 observations.
    colours = ["4472C4", "ED7D31", "C00000"]
    names = ["Setosa", "Versicolor", "Virginica"]
    for group, (name, colour) in enumerate(zip(names, colours)):
        first_row = 2 + group * 50
        last_row = first_row + 49
        x_values = Reference(data_sheet, min_col=3, min_row=first_row, max_row=last_row)
        y_values = Reference(data_sheet, min_col=4, min_row=first_row, max_row=last_row)
        series = Series(y_values, x_values, title=name)
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.marker.graphicalProperties.solidFill = colour
        series.marker.graphicalProperties.line.solidFill = colour
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
    summary.add_chart(chart, "D3")

    workbook.save(OUTPUT)


def verify(rows: list[list[object]]) -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    data_sheet = workbook["Iris Data"]
    assert data_sheet.max_row == len(rows) + 1
    assert data_sheet.max_column == 5
    assert len(data_sheet.tables) == 1
    assert len(workbook["Summary"]._charts) == 1
    assert Counter(data_sheet.cell(row, 5).value for row in range(2, data_sheet.max_row + 1)) == Counter(
        row[4] for row in rows
    )


if __name__ == "__main__":
    iris_rows = read_data()
    build_workbook(iris_rows)
    verify(iris_rows)
    print(f"Created and verified: {OUTPUT}")
    print(f"Rows: {len(iris_rows)} | Columns: {len(HEADERS)}")
