"""Create a polished three-sheet Excel workbook for Week 3 Activity 2."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HERE = Path(__file__).parent
SOURCES = (
    ("Cleaned Data", HERE.parent / "activity1" / "cleaned_dataset.csv"),
    ("Completed Data", HERE / "completed_dataset.csv"),
    ("Prediction Results", HERE / "prediction_results.csv"),
)
OUTPUT = HERE / "Week3_Activity2_Missing_Value_Prediction.xlsx"

NAVY = "17365D"
BLUE = "2F75B5"
PALE_BLUE = "D9EAF7"
PALE_GREEN = "E2F0D9"
PALE_AMBER = "FFF2CC"
PALE_RED = "FCE4D6"
WHITE = "FFFFFF"
GRID = "D9E2F3"


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        rows = list(reader)
    return rows[0], rows[1:]


def convert_value(header: str, value: str):
    if value == "":
        return None
    if header in {"ID", "Age", "Net worth", "Salary"}:
        return float(value)
    if header == "Join Date":
        return datetime.strptime(value, "%Y-%m-%d").date()
    return value


def style_sheet(sheet, title: str, subtitle: str, headers: list[str], rows: list[list[str]], table_name: str) -> None:
    column_count = len(headers)
    last_column = get_column_letter(column_count)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = sheet.cell(1, 1, title)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(color=WHITE, bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    subtitle_cell = sheet.cell(2, 1, subtitle)
    subtitle_cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
    subtitle_cell.font = Font(color=NAVY, italic=True, size=10)
    subtitle_cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 30

    for column, header in enumerate(headers, 1):
        cell = sheet.cell(3, column, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[3].height = 34

    for row_number, row in enumerate(rows, 4):
        for column, value in enumerate(row, 1):
            header = headers[column - 1]
            cell = sheet.cell(row_number, column, convert_value(header, value))
            cell.alignment = Alignment(vertical="top", wrap_text=header in {"Interpretation"})
            cell.border = Border(bottom=Side(style="hair", color=GRID))
            if row_number % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")

    last_row = 3 + len(rows)
    table = Table(displayName=table_name, ref=f"A3:{last_column}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A3:{last_column}{last_row}"

    for column_index, header in enumerate(headers, 1):
        values = [header] + [str(row[column_index - 1]) for row in rows]
        longest = max(len(value) for value in values)
        width = min(max(longest + 2, 11), 38)
        if header in {"Interpretation", "Prediction note", "Country prediction note"}:
            width = 38
        elif header.endswith("source"):
            width = 22
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    for column_index, header in enumerate(headers, 1):
        letter = get_column_letter(column_index)
        if header in {"Net worth", "Salary"}:
            for cell in sheet[f"{letter}4:{letter}{last_row}"]:
                cell[0].number_format = '#,##0.00'
        elif header in {"Age", "ID"}:
            for cell in sheet[f"{letter}4:{letter}{last_row}"]:
                cell[0].number_format = '0.##'
        elif header == "Join Date":
            for cell in sheet[f"{letter}4:{letter}{last_row}"]:
                cell[0].number_format = 'yyyy-mm-dd'


def add_source_highlighting(sheet) -> None:
    headers = {cell.value: cell.column for cell in sheet[3]}
    for header, column in headers.items():
        if not str(header).endswith(" source"):
            continue
        for row in range(4, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            value = str(cell.value or "")
            if value.startswith("Predicted"):
                cell.fill = PatternFill("solid", fgColor=PALE_AMBER)
                cell.font = Font(color="9C6500", bold=True)
            elif value == "Observed":
                cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
                cell.font = Font(color="375623")
            else:
                cell.fill = PatternFill("solid", fgColor=PALE_RED)
                cell.font = Font(color="9C0006")


def add_selection_highlighting(sheet) -> None:
    headers = {cell.value: cell.column for cell in sheet[3]}
    selected_column = headers["Selected model"]
    result_column = headers["Selected result"]
    for row in range(4, sheet.max_row + 1):
        for column in (selected_column, result_column):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
            cell.font = Font(color="375623", bold=True)


def build() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    subtitles = {
        "Cleaned Data": "Cleaned observations from Activity 1. Genuine unknown values remain blank.",
        "Completed Data": "Selected estimates are inserted for presentation; source columns identify every observed, predicted, or missing value.",
        "Prediction Results": "Linear and degree-2 polynomial results are compared using leave-one-out cross-validation. Lower RMSE selects the model.",
    }

    for index, (sheet_name, source) in enumerate(SOURCES, 1):
        headers, rows = read_csv(source)
        sheet = workbook.create_sheet(sheet_name)
        style_sheet(sheet, f"Week 3 Activity 2 — {sheet_name}", subtitles[sheet_name], headers, rows, f"Activity2Table{index}")
        if sheet_name == "Completed Data":
            add_source_highlighting(sheet)
        elif sheet_name == "Prediction Results":
            add_selection_highlighting(sheet)

    workbook.properties.title = "Week 3 Activity 2 — Missing Value Prediction"
    workbook.properties.subject = "Linear and polynomial regression comparison"
    workbook.properties.creator = "MSE803 Data Analytics"
    workbook.save(OUTPUT)


def verify() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    assert workbook.sheetnames == [name for name, _ in SOURCES]
    assert workbook["Cleaned Data"].max_row == 12
    assert workbook["Completed Data"].max_row == 12
    assert workbook["Prediction Results"].max_row == 9
    assert workbook["Completed Data"]["B11"].value == "Heidi"
    assert workbook["Prediction Results"]["A9"].value == "Heidi"
    for sheet in workbook.worksheets:
        assert len(sheet.tables) == 1
        assert sheet.freeze_panes == "A4"
    workbook.close()


if __name__ == "__main__":
    build()
    verify()
    print(OUTPUT)
